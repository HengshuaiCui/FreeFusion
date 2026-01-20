import numpy as np
from PIL import Image
# from Metric import *
from natsort import natsorted
from tqdm import tqdm
import os
import statistics
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
# from metric import *
warnings.filterwarnings("ignore")
from Evaluator import *

def write_excel(excel_name='metric.xlsx', worksheet_name='VIF', column_index=0, data=None):
    try:
        workbook = load_workbook(excel_name)
    except FileNotFoundError:
        workbook = Workbook()

    if worksheet_name in workbook.sheetnames:
        worksheet = workbook[worksheet_name]
    else:
        worksheet = workbook.create_sheet(title=worksheet_name)

    column = get_column_letter(column_index + 1)
    for i, value in enumerate(data):
        cell = worksheet[column + str(i + 1)]
        cell.value = value

    workbook.save(excel_name)


def evaluation_one(ir_name, vi_name, f_name):
    ir = image_read_cv2(os.path.join(ir_name), 'GRAY')
    vi = image_read_cv2(os.path.join(vi_name), 'GRAY')
    fi = image_read_cv2(os.path.join(f_name), 'GRAY')
    EN = Evaluator.EN(fi)
    SD = Evaluator.SD(fi)
    SF = Evaluator.SF(fi)
    AG = Evaluator.AG(fi)
    MI = Evaluator.MI(fi, ir, vi)
    SCD = Evaluator.SCD(fi, ir, vi)
    VIFF = Evaluator.VIFF(fi, ir, vi)
    Qabf = Evaluator.Qabf(fi, ir, vi)
    SSIM = Evaluator.SSIM(fi, ir, vi)

    return EN, SD, SF, AG, MI, SCD, VIFF, Qabf, SSIM


if __name__ == '__main__':
    with_mean = True
    dataset_name = 'mfnet'
    ir_dir = os.path.join('/mnt/disk/fusion_datasets/mfnet/test/ir')
    vi_dir = os.path.join('/mnt/disk/fusion_datasets/mfnet/test/rgb')
    models_list = []
    fus_dir = os.path.join('/mnt/disk/PythonProjects/CLIP_Fuison/MMFusionNet/results/mfnet/')#/
    save_dir = './metric_mfnet'
    os.makedirs(save_dir, exist_ok=True)
    methods = ['method_name']
    for method in methods:
        EN_list = []
        SD_list = []
        SF_list = []
        AG_list = []
        MI_list = []
        SCD_list = []
        VIFF_list = []
        Qabf_list = []
        SSIM_list = []
        filename_list = ['']
        fus_method_dir = os.path.join(fus_dir, method)
        metric_save_name = os.path.join(save_dir, 'metric_{}_{}.xlsx'.format(dataset_name, method))
        filelist = natsorted(os.listdir(fus_method_dir))
        eval_bar = tqdm(filelist)
        for _, item in enumerate(eval_bar):
            ir_name = os.path.join(ir_dir, item)
            vi_name = os.path.join(vi_dir, item)
            f_name = os.path.join(fus_method_dir, item)

            EN, SD, SF, AG, MI, SCD, VIFF, Qabf, SSIM = evaluation_one(ir_name, vi_name, f_name)
            EN_list.append(EN)
            SD_list.append(SD)
            SF_list.append(SF)
            AG_list.append(AG)
            MI_list.append(MI)
            SCD_list.append(SCD)
            VIFF_list.append(VIFF)
            Qabf_list.append(Qabf)
            SSIM_list.append(SSIM)
            filename_list.append(item)
            eval_bar.set_description("{} | {}".format(method, item))
        if with_mean:
            Avg_EN = np.mean(EN_list)
            Avg_SD = np.mean(SD_list)
            Avg_SF = np.mean(SF_list)
            Avg_AG = np.mean(AG_list)
            Avg_MI = np.mean(MI_list)
            Avg_SCD = np.mean(SCD_list)
            Avg_VIFF = np.mean(VIFF_list)
            Avg_Qabf = np.mean(Qabf_list)
            Avg_SSIM = np.mean(SSIM_list)

            EN_list.append(Avg_EN)
            SD_list.append(Avg_SD)
            SF_list.append(Avg_SF)
            AG_list.append(Avg_AG)
            MI_list.append(Avg_MI)
            SCD_list.append(Avg_SCD)
            VIFF_list.append(Avg_VIFF)
            Qabf_list.append(Avg_Qabf)
            SSIM_list.append(Avg_SSIM)
            filename_list.append('mean')

            print("Avg metric: \n  Avg_EN:%.4f, Avg_SD:%.4f, Avg_SF:%.4f, Avg_AG:%.4f, Avg_MI:%.4f, Avg_SCD:%.4f, Avg_VIFF:%.4f, Avg_Qabf:%.4f, "
                  "Avg_SSIM:%.4f"
                  %(Avg_EN, Avg_SD, Avg_SF, Avg_AG, Avg_MI, Avg_SCD, Avg_VIFF, Avg_Qabf, Avg_SSIM))

        ## 保留三位小数
        EN_list = [round(x, 3) for x in EN_list]
        SD_list = [round(x, 3) for x in SD_list]
        SF_list = [round(x, 3) for x in SF_list]
        AG_list = [round(x, 3) for x in AG_list]
        MI_list = [round(x, 3) for x in MI_list]
        SCD_list = [round(x, 3) for x in SCD_list]
        VIFF_list = [round(x, 3) for x in VIFF_list]
        Qabf_list = [round(x, 3) for x in Qabf_list]
        SSIM_list = [round(x, 3) for x in SSIM_list]

        EN_list.insert(0, '{}'.format(method))
        SD_list.insert(0, '{}'.format(method))
        SF_list.insert(0, '{}'.format(method))
        AG_list.insert(0, '{}'.format(method))
        MI_list.insert(0, '{}'.format(method))
        SCD_list.insert(0, '{}'.format(method))
        VIFF_list.insert(0, '{}'.format(method))
        Qabf_list.insert(0, '{}'.format(method))
        SSIM_list.insert(0, '{}'.format(method))

        write_excel(metric_save_name, 'EN', 0, filename_list)
        write_excel(metric_save_name, "SD", 0, filename_list)
        write_excel(metric_save_name, "SF", 0, filename_list)
        write_excel(metric_save_name, "AG", 0, filename_list)
        write_excel(metric_save_name, "MI", 0, filename_list)
        write_excel(metric_save_name, "SCD", 0, filename_list)
        write_excel(metric_save_name, "VIFF", 0, filename_list)
        write_excel(metric_save_name, "Qabf", 0, filename_list)
        write_excel(metric_save_name, "SSIM", 0, filename_list)

        write_excel(metric_save_name, 'EN', 1, EN_list)
        write_excel(metric_save_name, "SD", 1, SD_list)
        write_excel(metric_save_name, "SF", 1, SF_list)
        write_excel(metric_save_name, "AG", 1, AG_list)
        write_excel(metric_save_name, "MI", 1, MI_list)
        write_excel(metric_save_name, "SCD", 1, SCD_list)
        write_excel(metric_save_name, "VIFF", 1, VIFF_list)
        write_excel(metric_save_name, "Qabf", 1, Qabf_list)
        write_excel(metric_save_name, "SSIM", 1, SSIM_list)